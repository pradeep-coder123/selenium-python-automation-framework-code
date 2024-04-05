import inspect
import logging
import softest
from openpyxl import Workbook, load_workbook
import csv


class Utils(softest.TestCase):

    def assertListitemText(self,list,value):
        for stop in list:
            print("The text is:" +stop.text)
            # assert stop.text == value
            # print("assert pass")
            self.soft_assert(self.assertEqual, stop.text, value)

            if stop.text == value:
                print("text passed")

            else:
                print("text failed")

        self.assert_all()



    def custome_logger(logLevel = logging.DEBUG):
        # set class/method name from where its called
        # logger_name = inspect.stack()[1][3]

        # create logger
        logger = logging.getLogger(__name__) # to create an object i.e. logger object, it will help to set an object
        logger.setLevel(logLevel) # set the level of the object, it will log all the messsags from debug and above

        # create console handler = you want to display your logs on the console you configure the console handler or
        # file handler = move your logs to the files or the log files you configure the file handler
        # and set level to debug
        # ch.setLevel(logging.DEBUG)
        # fh = logging.FileHandler("automation.log")
        fh = logging.FileHandler("C:\\Users\\pmondal\\PycharmProjects\\pythonProject\\PythonSeleniumproject1\\TestFramework\\utilities\\automation.log", mode='w')
        # create formatter - how you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')

        # add formatter to ch -> add formatter to console or file handler
        fh.setFormatter(formatter)

        # add ch= console handler to logger because logger is the main things which logs the messages
        logger.addHandler(fh)

        return logger


    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename = file_name)
        sh = wb[sheet] # if you want to provide the specific sheet.

        row_ct = sh.max_row
        col_ct = sh.max_column


        for i in range(2, row_ct+1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row = i, column= j).value)
            datalist.append(row)
        return datalist


    def read_data_from_csv(filename):
       #Create an empty List
       dataList = []

       #Open CSV file
       csvdata = open(filename,"r")

       #Create CSV reader
       reader = csv.reader(csvdata)

       #Skip header
       next(reader)

       #Add CSV rows to list
       for rows in reader:
           dataList.append(rows)

       return dataList